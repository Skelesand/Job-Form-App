import datetime
import os
import time
from flask import Flask, jsonify, render_template, request, redirect, url_for
import openpyxl

app = Flask(__name__)

EXCEL_PATH = os.getenv(
    "EXCEL_PATH",
    r"data\Scan_Log_Dataset.xlsx"
)


def get_existing_projects(filepath):
    """Reads the Excel sheet and returns a unique list of project names from Column A."""
    projects = []
    try:
        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2, max_col=1):
                if row[0].value:
                    val = str(row[0].value).strip()
                    if val and val not in projects:
                        projects.append(val)
            wb.close()
    except Exception as e:
        print(f"[Warning] Could not read project list: {e}")
    return sorted(projects)


def get_all_projects(filepath):
    """Reads the entire Excel sheet and returns a list of dictionaries for all logged jobs."""
    projects = []
    try:
        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb.active
            for row in sheet.iter_rows(min_row=2):
                if row[0].value:
                    projects.append({
                        "project_name": str(row[0].value).strip(),
                        "project_type": row[1].value if len(row) > 1 and row[1].value else "",
                        "sector": row[2].value if len(row) > 2 and row[2].value else "",
                        "sqft": row[3].value if len(row) > 3 and row[3].value else "",
                        "levels": row[4].value if len(row) > 4 and row[4].value else "",
                        "partition_density": row[5].value if len(row) > 5 and row[5].value else "",
                        "site_condition": row[6].value if len(row) > 6 and row[6].value else "",
                        "interior": row[7].value if len(row) > 7 and row[7].value else "",
                        "exterior": row[8].value if len(row) > 8 and row[8].value else "",
                        "roof": row[9].value if len(row) > 9 and row[9].value else "",
                        "coverage": row[10].value if len(row) > 10 and row[10].value else "",
                        "scan_count": row[11].value if len(row) > 11 and row[11].value else "",
                        "timestamp": row[12].value if len(row) > 12 and row[12].value else ""
                    })
            wb.close()
    except Exception as e:
        print(f"[Error] Could not read all projects: {e}")
    return projects


def get_project_details(filepath, project_name):
    """Searches for a project by name and returns its row values as a dictionary."""
    try:
        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = wb.active

            for row in sheet.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).strip() == project_name.strip():
                    data = {
                        "project_type": row[1].value if row[1].value else "",
                        "sector": row[2].value if row[2].value else "",
                        "sqft": row[3].value if row[3].value else "",
                        "levels": row[4].value if row[4].value else "",
                        "partition_density": row[5].value if row[5].value else "",
                        "site_condition": row[6].value if row[6].value else "",
                        "interior": row[7].value if row[7].value else "",
                        "exterior": row[8].value if row[8].value else "",
                        "roof": row[9].value if row[9].value else "",
                        "coverage": row[10].value if row[10].value else "",
                        "scan_count": row[11].value if row[11].value else "",
                    }
                    wb.close()
                    return data
            wb.close()
    except Exception as e:
        print(f"[Error] Failed to fetch project details: {e}")
    return None


def update_or_append_to_excel(filepath, data_row):
    max_retries = 3
    retry_delay = 2

    for attempt in range(max_retries):
        try:
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active

            project_name = data_row[0]
            row_updated = False

            for row in sheet.iter_rows(min_row=1, max_col=1):
                if row[0].value == project_name:
                    current_row_idx = row[0].row
                    for col_idx, value in enumerate(data_row, start=1):
                        sheet.cell(row=current_row_idx, column=col_idx, value=value)
                    row_updated = True
                    print(f"[Info] Updated existing row for project: {project_name}")
                    break

            if not row_updated:
                sheet.append(data_row)
                print(f"[Info] Appended new row for project: {project_name}")

            wb.save(filepath)
            wb.close()
            return True

        except PermissionError:
            print(f"[Warning] File locked. Retrying in {retry_delay} seconds... (Attempt {attempt + 1}/{max_retries})")
            time.sleep(retry_delay)

    print("[Error] Could not write to Excel. File locked.")
    return False


@app.route("/")
def home():
    """Dashboard view showing all logged jobs."""
    all_jobs = get_all_projects(EXCEL_PATH)
    return render_template("index.html", jobs=all_jobs)


@app.route("/form")
def form():
    """Data entry view for logging/editing a project."""
    project_list = get_existing_projects(EXCEL_PATH)
    return render_template("form.html", projects=project_list)


@app.route("/get_project_data")
def get_project_data():
    project_name = request.args.get("name")
    if not project_name:
        return jsonify({"success": False, "error": "No project name provided"})

    data = get_project_details(EXCEL_PATH, project_name)
    if data:
        return jsonify({"success": True, "data": data})
    return jsonify({"success": False, "error": "Project not found"})


def remove_project_from_excel(filepath, project_name):
    """Remove a project row from Excel by project name."""
    max_retries = 3
    retry_delay = 2

    for attempt in range(max_retries):
        try:
            wb = openpyxl.load_workbook(filepath)
            sheet = wb.active
            
            # Find and delete the row with matching project name
            for row in sheet.iter_rows(min_row=2):
                if row[0].value and str(row[0].value).strip() == project_name.strip():
                    sheet.delete_rows(row[0].row)
                    wb.save(filepath)
                    wb.close()
                    print(f"[Info] Deleted row for project: {project_name}")
                    return True
            
            wb.close()
            # If we get here, project wasn't found
            return False

        except PermissionError:
            print(f"[Warning] File locked. Retrying in {retry_delay} seconds... (Attempt {attempt + 1}/{max_retries})")
            time.sleep(retry_delay)
    
    print("[Error] Could not write to Excel. File locked.")
    return False


@app.route("/delete", methods=["POST"])
def delete_project():
    try:
        data = request.get_json()
        project_name = data.get("project_name")
        
        if not project_name:
            return jsonify({"success": False, "error": "No project name provided"})
        
        # Remove the project from Excel
        success = remove_project_from_excel(EXCEL_PATH, project_name)
        
        if success:
            return jsonify({"success": True, "message": "Project deleted successfully"})
        else:
            return jsonify({"success": False, "error": "Failed to delete project"})
    except Exception as e:
        print(f"[Error] Error deleting project: {e}")
        return jsonify({"success": False, "error": str(e)})


@app.route("/submit", methods=["POST"])
def submit():
    try:
        project_name = request.form.get("project_name")
        project_type = request.form.get("project_type")
        sector = request.form.get("sector")
        sqft = request.form.get("sqft")
        levels = request.form.get("levels")
        partition_density = request.form.get("partition_density")
        site_condition = request.form.get("site_condition")
        interior = request.form.get("interior")
        exterior = request.form.get("exterior")
        roof = request.form.get("roof")
        coverage = request.form.get("coverage")
        scan_count = request.form.get("scan_count")
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Validate required fields
        if not project_name or not project_name.strip():
            return "<h3>Error: Project name is required.</h3><a href='/form'>Go Back</a>"

        # Validate and convert numeric fields with error handling
        try:
            sqft = int(sqft)
            levels = int(levels)
            partition_density = int(partition_density)
            site_condition = int(site_condition)
            coverage = float(coverage)
            scan_count = int(scan_count)
        except (ValueError, TypeError) as e:
            return f"<h3>Error: Invalid numeric input. {str(e)}</h3><a href='/form'>Go Back</a>"

        new_row = [
            project_name.strip(),
            project_type,
            sector,
            sqft,
            levels,
            partition_density,
            site_condition,
            interior,
            exterior,
            roof,
            coverage,
            scan_count,
            timestamp,
        ]

        success = update_or_append_to_excel(EXCEL_PATH, new_row)

        if success:
            return "<h3>Job logged/updated successfully! Box will sync the file shortly.</h3><a href='/'>Go to Home Dashboard</a>"
        else:
            return "<h3>Error: The Excel file is currently open/locked. Please try submitting again.</h3><a href='/form'>Go Back</a>"
    except Exception as e:
        print(f"[Error] Unexpected error in /submit: {e}")
        return f"<h3>Error: An unexpected error occurred. Please try again.</h3><a href='/form'>Go Back</a>"


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)